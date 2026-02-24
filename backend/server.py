from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import json

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ── Models ──────────────────────────────────────────────────────────────
class ChatMessage(BaseModel):
    role: str
    content: str
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = "default"

class ChatResponse(BaseModel):
    reply: str
    session_id: str

# ── Data Seeding ────────────────────────────────────────────────────────
async def seed_data():
    count = await db.erw_samples.count_documents({})
    if count > 0:
        logger.info(f"Database already seeded with {count} samples")
        return

    logger.info("Seeding database from Excel file...")
    try:
        import openpyxl
        xlsx_path = ROOT_DIR.parent / 'das1.xlsx'
        if not xlsx_path.exists():
            logger.warning("das1.xlsx not found, skipping seed")
            return

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        # Parse ERW Results
        ws = wb['ERW Results']
        headers_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        clean_headers = []
        for h in headers_row:
            if h is None:
                clean_headers.append(None)
            else:
                clean_headers.append(str(h).split('\n')[0].strip())

        samples = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row[0] or str(row[0]).strip() == '':
                continue
            # Skip section headers
            if row[2] is None and row[3] is None and row[4] is None:
                continue

            sample = {
                "id": str(uuid.uuid4()),
                "feedstock": "calcite",
                "omega_threshold": 5,
                "sample_no": str(row[0]) if row[0] else "",
                "river_type": str(row[1]) if row[1] else "",
                "latitude": float(row[2]) if row[2] else None,
                "longitude": float(row[3]) if row[3] else None,
                "ph": float(row[4]) if row[4] else None,
                "alkalinity": float(row[5]) if row[5] else None,
                "temp_c": float(row[6]) if row[6] else None,
                "ca": float(row[7]) if row[7] else None,
                "mg": float(row[8]) if row[8] else None,
                "na": float(row[9]) if row[9] else None,
                "k": float(row[10]) if row[10] else None,
                "cl": float(row[11]) if row[11] else None,
                "so4": float(row[12]) if row[12] else None,
                "no3": float(row[13]) if row[13] else None,
                "salinity": float(row[15]) if row[15] else None,
                "ksp": float(row[17]) if row[17] else None,
                "hco3": float(row[20]) if row[20] else None,
                "co3": float(row[21]) if row[21] else None,
                "co2_aq": float(row[22]) if row[22] else None,
                "dic": float(row[23]) if row[23] else None,
                "pco2": float(row[24]) if row[24] else None,
                "omega_calcite": float(row[31]) if row[31] else None,
                "si_calcite": float(row[32]) if row[32] else None,
                "state": str(row[34]) if row[34] else "",
                "region": str(row[35]) if row[35] else "",
                "river_name": str(row[36]) if row[36] else "",
                "discharge": float(row[37]) if row[37] else None,
                "source": str(row[38]) if row[38] else "",
                # ERW model outputs
                "j_steps": int(row[40]) if row[40] is not None else None,
                "k_steps": int(row[41]) if row[41] is not None else None,
                "rock_addition": float(row[42]) if row[42] else None,
                "omega_flag": int(row[43]) if row[43] is not None else None,
                "success_flag": int(row[44]) if row[44] is not None else None,
                "omega_final": float(row[47]) if row[47] else None,
                "ca_final": float(row[48]) if row[48] else None,
                "alk_final": float(row[50]) if row[50] else None,
                "dic_final": float(row[51]) if row[51] else None,
                "ph_final": float(row[53]) if row[53] else None,
                "pco2_final": float(row[55]) if row[55] else None,
                "discharge_ms": float(row[57]) if row[57] else None,
                "cdr_mol_s": float(row[58]) if row[58] else None,
                "cdr_t_yr": float(row[59]) if row[59] else None,
                "cdr_kt_yr": float(row[60]) if row[60] else None,
            }
            samples.append(sample)

        if samples:
            await db.erw_samples.insert_many(samples)
            logger.info(f"Seeded {len(samples)} ERW samples")

        # Parse Summary Statistics
        ws2 = wb['Summary Statistics']
        summaries = []
        for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, values_only=True):
            if not row[0] or str(row[0]).strip() == '':
                continue
            summary = {
                "id": str(uuid.uuid4()),
                "feedstock": "calcite",
                "omega_threshold": 5,
                "region": str(row[0]),
                "add_mean": float(row[1]) if row[1] else 0,
                "add_median": float(row[2]) if row[2] else 0,
                "add_std": float(row[3]) if row[3] else 0,
                "add_min": float(row[4]) if row[4] else 0,
                "add_max": float(row[5]) if row[5] else 0,
                "n_samples": int(row[6]) if row[6] else 0,
                "omega_mean": float(row[7]) if row[7] else 0,
                "omega_median": float(row[8]) if row[8] else 0,
                "omega_std": float(row[9]) if row[9] else 0,
                "cdr_mean": float(row[10]) if row[10] else 0,
                "cdr_total": float(row[11]) if row[11] else 0,
                "n_with_q": int(row[12]) if row[12] else 0,
                "success_pct": float(row[13]) if row[13] else 0,
            }
            summaries.append(summary)

        if summaries:
            await db.summary_stats.insert_many(summaries)
            logger.info(f"Seeded {len(summaries)} summary stats")

        # Create default feedstock entry
        await db.feedstocks.insert_one({
            "id": str(uuid.uuid4()),
            "name": "calcite",
            "omega_thresholds": [5],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "sample_count": len(samples),
        })
        logger.info("Seeding complete")

    except Exception as e:
        logger.error(f"Seeding error: {e}")
        import traceback
        traceback.print_exc()


@app.on_event("startup")
async def startup():
    await seed_data()


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ── Dashboard Overview ──────────────────────────────────────────────────
@api_router.get("/dashboard/overview")
async def dashboard_overview(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "cdr_t_yr": {"$ne": None, "$gt": 0}}},
        {"$group": {
            "_id": None,
            "total_cdr_t_yr": {"$sum": "$cdr_t_yr"},
            "avg_cdr_t_yr": {"$avg": "$cdr_t_yr"},
            "total_samples": {"$sum": 1},
            "avg_ph": {"$avg": "$ph"},
            "avg_alkalinity": {"$avg": "$alkalinity"},
            "avg_rock_addition": {"$avg": "$rock_addition"},
            "avg_omega_final": {"$avg": "$omega_final"},
        }}
    ]
    result = await db.erw_samples.aggregate(pipeline).to_list(1)

    total_samples = await db.erw_samples.count_documents({"feedstock": feedstock, "omega_threshold": omega})
    successful = await db.erw_samples.count_documents({"feedstock": feedstock, "omega_threshold": omega, "success_flag": 1})

    if result:
        r = result[0]
        return {
            "total_cdr_t_yr": r.get("total_cdr_t_yr", 0),
            "avg_cdr_t_yr": r.get("avg_cdr_t_yr", 0),
            "total_samples": total_samples,
            "samples_with_cdr": r.get("total_samples", 0),
            "avg_ph": r.get("avg_ph", 0),
            "avg_alkalinity": r.get("avg_alkalinity", 0),
            "avg_rock_addition": r.get("avg_rock_addition", 0),
            "avg_omega_final": r.get("avg_omega_final", 0),
            "success_rate": (successful / total_samples * 100) if total_samples > 0 else 0,
            "feedstock": feedstock,
            "omega_threshold": omega,
        }
    return {"total_cdr_t_yr": 0, "avg_cdr_t_yr": 0, "total_samples": total_samples, "samples_with_cdr": 0,
            "avg_ph": 0, "avg_alkalinity": 0, "avg_rock_addition": 0, "avg_omega_final": 0,
            "success_rate": 0, "feedstock": feedstock, "omega_threshold": omega}


# ── Region-wise Summary ─────────────────────────────────────────────────
@api_router.get("/summary")
async def get_summary(feedstock: str = "calcite", omega: int = 5):
    docs = await db.summary_stats.find(
        {"feedstock": feedstock, "omega_threshold": omega}, {"_id": 0}
    ).to_list(100)
    return docs


# ── Region-wise aggregation from samples ─────────────────────────────────
@api_router.get("/regions/cdr")
async def regions_cdr(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "region": {"$ne": ""}}},
        {"$group": {
            "_id": "$region",
            "total_cdr": {"$sum": "$cdr_t_yr"},
            "avg_cdr": {"$avg": "$cdr_t_yr"},
            "count": {"$sum": 1},
            "avg_ph": {"$avg": "$ph"},
            "avg_rock_add": {"$avg": "$rock_addition"},
        }},
        {"$sort": {"total_cdr": -1}}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(50)
    return [{"region": r["_id"], "total_cdr": r["total_cdr"] or 0, "avg_cdr": r["avg_cdr"] or 0,
             "count": r["count"], "avg_ph": r["avg_ph"] or 0, "avg_rock_add": r["avg_rock_add"] or 0} for r in results]


# ── Samples endpoint ─────────────────────────────────────────────────────
@api_router.get("/samples")
async def get_samples(feedstock: str = "calcite", omega: int = 5, region: Optional[str] = None,
                      state: Optional[str] = None, limit: int = 200, skip: int = 0):
    query = {"feedstock": feedstock, "omega_threshold": omega}
    if region:
        query["region"] = region
    if state:
        query["state"] = state
    docs = await db.erw_samples.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.erw_samples.count_documents(query)
    return {"samples": docs, "total": total}


# ── Map data ──────────────────────────────────────────────────────────────
@api_router.get("/samples/map")
async def get_map_data(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega,
                     "latitude": {"$ne": None}, "longitude": {"$ne": None}}},
        {"$project": {"_id": 0, "latitude": 1, "longitude": 1, "river_name": 1,
                      "state": 1, "region": 1, "cdr_t_yr": 1, "cdr_kt_yr": 1,
                      "ph": 1, "rock_addition": 1, "omega_final": 1, "sample_no": 1}}
    ]
    docs = await db.erw_samples.aggregate(pipeline).to_list(2000)
    return docs


# ── Filters (regions, states, rivers) ─────────────────────────────────────
@api_router.get("/filters")
async def get_filters(feedstock: str = "calcite", omega: int = 5):
    query = {"feedstock": feedstock, "omega_threshold": omega}
    regions = await db.erw_samples.distinct("region", query)
    states = await db.erw_samples.distinct("state", query)
    return {
        "regions": sorted([r for r in regions if r]),
        "states": sorted([s for s in states if s]),
    }


# ── Analytics data ────────────────────────────────────────────────────────
@api_router.get("/analytics/distributions")
async def get_distributions(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega}},
        {"$project": {"_id": 0, "ph": 1, "alkalinity": 1, "dic": 1, "pco2": 1,
                      "temp_c": 1, "rock_addition": 1, "cdr_t_yr": 1, "omega_calcite": 1,
                      "omega_final": 1, "discharge": 1, "region": 1, "state": 1,
                      "ca": 1, "mg": 1, "hco3": 1, "salinity": 1}}
    ]
    docs = await db.erw_samples.aggregate(pipeline).to_list(2000)
    return docs


# ── Scatter data ──────────────────────────────────────────────────────────
@api_router.get("/analytics/scatter")
async def get_scatter(feedstock: str = "calcite", omega: int = 5, x_field: str = "discharge",
                      y_field: str = "cdr_t_yr"):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega,
                     x_field: {"$ne": None}, y_field: {"$ne": None}}},
        {"$project": {"_id": 0, x_field: 1, y_field: 1, "region": 1, "river_name": 1, "state": 1}}
    ]
    docs = await db.erw_samples.aggregate(pipeline).to_list(2000)
    return docs


# ── Feedstock management ──────────────────────────────────────────────────
@api_router.get("/feedstocks")
async def list_feedstocks():
    docs = await db.feedstocks.find({}, {"_id": 0}).to_list(50)
    return docs


@api_router.post("/feedstock/upload")
async def upload_feedstock(file: UploadFile = File(...), feedstock_name: str = "unknown",
                           omega_threshold: int = 5):
    try:
        import openpyxl
        import io
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        ws = wb.active
        headers_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

        samples = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row[0] or str(row[0]).strip() == '':
                continue
            if row[2] is None and row[3] is None and row[4] is None:
                continue

            sample = {
                "id": str(uuid.uuid4()),
                "feedstock": feedstock_name.lower(),
                "omega_threshold": omega_threshold,
                "sample_no": str(row[0]) if row[0] else "",
                "river_type": str(row[1]) if row[1] else "",
                "latitude": float(row[2]) if row[2] else None,
                "longitude": float(row[3]) if row[3] else None,
                "ph": float(row[4]) if row[4] else None,
                "alkalinity": float(row[5]) if row[5] else None,
                "temp_c": float(row[6]) if row[6] else None,
                "ca": float(row[7]) if row[7] else None,
                "mg": float(row[8]) if row[8] else None,
                "na": float(row[9]) if row[9] else None,
                "k": float(row[10]) if row[10] else None,
                "cl": float(row[11]) if row[11] else None,
                "so4": float(row[12]) if row[12] else None,
                "no3": float(row[13]) if row[13] else None,
                "salinity": float(row[15]) if row[15] else None,
                "ksp": float(row[17]) if row[17] else None,
                "hco3": float(row[20]) if row[20] else None,
                "co3": float(row[21]) if row[21] else None,
                "co2_aq": float(row[22]) if row[22] else None,
                "dic": float(row[23]) if row[23] else None,
                "pco2": float(row[24]) if row[24] else None,
                "omega_calcite": float(row[31]) if row[31] else None,
                "si_calcite": float(row[32]) if row[32] else None,
                "state": str(row[34]) if row[34] else "",
                "region": str(row[35]) if row[35] else "",
                "river_name": str(row[36]) if row[36] else "",
                "discharge": float(row[37]) if row[37] else None,
                "source": str(row[38]) if row[38] else "",
                "j_steps": int(row[40]) if row[40] is not None else None,
                "k_steps": int(row[41]) if row[41] is not None else None,
                "rock_addition": float(row[42]) if row[42] else None,
                "omega_flag": int(row[43]) if row[43] is not None else None,
                "success_flag": int(row[44]) if row[44] is not None else None,
                "omega_final": float(row[47]) if row[47] else None,
                "ca_final": float(row[48]) if row[48] else None,
                "alk_final": float(row[50]) if row[50] else None,
                "dic_final": float(row[51]) if row[51] else None,
                "ph_final": float(row[53]) if row[53] else None,
                "pco2_final": float(row[55]) if row[55] else None,
                "discharge_ms": float(row[57]) if row[57] else None,
                "cdr_mol_s": float(row[58]) if row[58] else None,
                "cdr_t_yr": float(row[59]) if row[59] else None,
                "cdr_kt_yr": float(row[60]) if row[60] else None,
            }
            samples.append(sample)

        if samples:
            await db.erw_samples.insert_many(samples)

        # Parse summary stats if present
        if 'Summary Statistics' in wb.sheetnames:
            ws2 = wb['Summary Statistics']
            summaries = []
            for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, values_only=True):
                if not row[0] or str(row[0]).strip() == '':
                    continue
                summary = {
                    "id": str(uuid.uuid4()),
                    "feedstock": feedstock_name.lower(),
                    "omega_threshold": omega_threshold,
                    "region": str(row[0]),
                    "add_mean": float(row[1]) if row[1] else 0,
                    "add_median": float(row[2]) if row[2] else 0,
                    "add_std": float(row[3]) if row[3] else 0,
                    "add_min": float(row[4]) if row[4] else 0,
                    "add_max": float(row[5]) if row[5] else 0,
                    "n_samples": int(row[6]) if row[6] else 0,
                    "omega_mean": float(row[7]) if row[7] else 0,
                    "omega_median": float(row[8]) if row[8] else 0,
                    "omega_std": float(row[9]) if row[9] else 0,
                    "cdr_mean": float(row[10]) if row[10] else 0,
                    "cdr_total": float(row[11]) if row[11] else 0,
                    "n_with_q": int(row[12]) if row[12] else 0,
                    "success_pct": float(row[13]) if row[13] else 0,
                }
                summaries.append(summary)
            if summaries:
                await db.summary_stats.insert_many(summaries)

        # Update or create feedstock entry
        existing = await db.feedstocks.find_one({"name": feedstock_name.lower()})
        if existing:
            thresholds = existing.get("omega_thresholds", [])
            if omega_threshold not in thresholds:
                thresholds.append(omega_threshold)
            await db.feedstocks.update_one(
                {"name": feedstock_name.lower()},
                {"$set": {"omega_thresholds": thresholds, "sample_count": existing.get("sample_count", 0) + len(samples)}}
            )
        else:
            await db.feedstocks.insert_one({
                "id": str(uuid.uuid4()),
                "name": feedstock_name.lower(),
                "omega_thresholds": [omega_threshold],
                "created_at": datetime.now(timezone.utc).isoformat(),
                "sample_count": len(samples),
            })

        return {"message": f"Uploaded {len(samples)} samples for {feedstock_name} at omega={omega_threshold}",
                "samples_count": len(samples)}
    except Exception as e:
        logger.error(f"Upload error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ── Omega Comparison ──────────────────────────────────────────────────────
@api_router.get("/comparison")
async def omega_comparison(feedstock: str = "calcite"):
    thresholds = await db.erw_samples.distinct("omega_threshold", {"feedstock": feedstock})
    results = []
    for omega in sorted(thresholds):
        pipeline = [
            {"$match": {"feedstock": feedstock, "omega_threshold": omega}},
            {"$group": {
                "_id": "$region",
                "total_cdr": {"$sum": "$cdr_t_yr"},
                "avg_cdr": {"$avg": "$cdr_t_yr"},
                "avg_rock_add": {"$avg": "$rock_addition"},
                "avg_ph_final": {"$avg": "$ph_final"},
                "avg_omega_final": {"$avg": "$omega_final"},
                "count": {"$sum": 1},
                "successful": {"$sum": {"$cond": [{"$eq": ["$success_flag", 1]}, 1, 0]}},
            }},
            {"$sort": {"total_cdr": -1}}
        ]
        regions = await db.erw_samples.aggregate(pipeline).to_list(50)

        total_pipeline = [
            {"$match": {"feedstock": feedstock, "omega_threshold": omega}},
            {"$group": {
                "_id": None,
                "total_cdr": {"$sum": "$cdr_t_yr"},
                "avg_rock_add": {"$avg": "$rock_addition"},
                "count": {"$sum": 1},
            }}
        ]
        totals = await db.erw_samples.aggregate(total_pipeline).to_list(1)
        total_info = totals[0] if totals else {"total_cdr": 0, "avg_rock_add": 0, "count": 0}

        results.append({
            "omega_threshold": omega,
            "total_cdr": total_info.get("total_cdr", 0) or 0,
            "avg_rock_add": total_info.get("avg_rock_add", 0) or 0,
            "total_samples": total_info.get("count", 0),
            "regions": [{"region": r["_id"], "total_cdr": r["total_cdr"] or 0,
                         "avg_cdr": r["avg_cdr"] or 0, "avg_rock_add": r["avg_rock_add"] or 0,
                         "count": r["count"],
                         "success_rate": (r["successful"] / r["count"] * 100) if r["count"] > 0 else 0}
                        for r in regions if r["_id"]]
        })
    return results


# ── AI Chat ───────────────────────────────────────────────────────────────
@api_router.post("/chat", response_model=ChatResponse)
async def chat_endpoint(req: ChatRequest):
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage

        api_key = os.environ.get("EMERGENT_LLM_KEY", "")

        # Build context from data
        summary_docs = await db.summary_stats.find({}, {"_id": 0}).to_list(100)
        feedstock_docs = await db.feedstocks.find({}, {"_id": 0}).to_list(50)

        # Get sample stats
        total = await db.erw_samples.count_documents({})
        regions = await db.erw_samples.distinct("region")
        states = await db.erw_samples.distinct("state")

        data_context = f"""You are an expert in Enhanced Rock Weathering (ERW) and carbon dioxide removal (CDR) for Indian rivers.

Available data summary:
- Total samples: {total}
- Regions: {', '.join([r for r in regions if r])}
- States: {', '.join([s for s in states if s])}
- Available feedstocks: {json.dumps([d.get('name') for d in feedstock_docs])}

Summary statistics by region:
{json.dumps(summary_docs, indent=2, default=str)}

Key definitions:
- ERW: Enhanced Rock Weathering - adding crusite/silicate minerals to rivers to capture CO2
- CDR: Carbon Dioxide Removal (measured in t CO2/yr)
- Omega (Ω): Calcite saturation index - higher values mean more saturated
- Rock addition: Amount of rock added (mol/kg)
- DIC: Dissolved Inorganic Carbon
- TA/Alkalinity: Total Alkalinity
- pCO2: Partial pressure of CO2

Answer questions about this data accurately. Use specific numbers from the data when possible. If you don't have data for a question, say so clearly."""

        # Get recent chat history
        history = await db.chat_messages.find(
            {"session_id": req.session_id}, {"_id": 0}
        ).sort("timestamp", 1).to_list(20)

        chat = LlmChat(
            api_key=api_key,
            session_id=f"erw_{req.session_id}",
            system_message=data_context
        )
        chat.with_model("openai", "gpt-5.2")

        # Add history messages
        for msg in history[-10:]:
            if msg["role"] == "user":
                await chat.send_message(UserMessage(text=msg["content"]))
            # assistant messages are already in the chat context

        response = await chat.send_message(UserMessage(text=req.message))

        # Store messages
        now = datetime.now(timezone.utc).isoformat()
        await db.chat_messages.insert_many([
            {"session_id": req.session_id, "role": "user", "content": req.message, "timestamp": now},
            {"session_id": req.session_id, "role": "assistant", "content": response, "timestamp": now},
        ])

        return ChatResponse(reply=response, session_id=req.session_id)

    except Exception as e:
        logger.error(f"Chat error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/chat/history")
async def chat_history(session_id: str = "default", limit: int = 50):
    docs = await db.chat_messages.find(
        {"session_id": session_id}, {"_id": 0}
    ).sort("timestamp", 1).to_list(limit)
    return docs


# ── State-wise aggregation ────────────────────────────────────────────────
@api_router.get("/states/cdr")
async def states_cdr(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "state": {"$ne": ""}}},
        {"$group": {
            "_id": "$state",
            "total_cdr": {"$sum": "$cdr_t_yr"},
            "avg_cdr": {"$avg": "$cdr_t_yr"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"total_cdr": -1}}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(50)
    return [{"state": r["_id"], "total_cdr": r["total_cdr"] or 0,
             "avg_cdr": r["avg_cdr"] or 0, "count": r["count"]} for r in results]


# ── Top Rivers ────────────────────────────────────────────────────────────
@api_router.get("/rivers/top")
async def top_rivers(feedstock: str = "calcite", omega: int = 5, limit: int = 20):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega,
                     "river_name": {"$ne": ""}, "cdr_t_yr": {"$ne": None}}},
        {"$group": {
            "_id": "$river_name",
            "total_cdr": {"$sum": "$cdr_t_yr"},
            "avg_cdr": {"$avg": "$cdr_t_yr"},
            "count": {"$sum": 1},
            "region": {"$first": "$region"},
            "state": {"$first": "$state"},
        }},
        {"$sort": {"total_cdr": -1}},
        {"$limit": limit}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(limit)
    return [{"river": r["_id"], "total_cdr": r["total_cdr"] or 0, "avg_cdr": r["avg_cdr"] or 0,
             "count": r["count"], "region": r.get("region", ""), "state": r.get("state", "")} for r in results]


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
